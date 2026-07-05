"""
SaaS 料金プラン (Excel) 生成テンプレート — saas-pricing-design スキル同梱

案件リポジトリの scripts/ にコピーし、`# TODO:` マーカー箇所を案件に合わせて調整して使う。
そのまま実行してもサンプル値で 料金プラン.xlsx を生成できる (スキルの自己テスト用)。

前提モデル (調整前のサンプル):
  - 1 契約 = 管理組織 1 + 利用テナント N 社 (N は 1 でも成立する料金設計)
  - 収入: 管理組織基本料 + テナント基本料 × 社数 + 従量 (込み人数超過分 × 単価)
  - コスト: 人件費 (単価 × 稼働時間) + 認証 (Cognito) + Cloudflare (Workers/D1) + R2 アーカイブ
  - 蓄積は人月ベース (平均 × 通常月 + 繁忙 × 繁忙月) で計算する

Numbers 互換ルール (references/excel-recipe.md 参照。違反すると全セル 0 になる):
  - シート間参照はクォートなし・相対参照 (=料金プラン!B19)
  - シート名に記号を使わない / 「=」始まりのテキスト禁止 / fullCalcOnLoad 必須

Usage:
    python3 generate_pricing_template.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

# TODO: 案件リポジトリでの出力先に変更する (例: docs/料金プラン.xlsx)
OUTPUT_PATH = Path.cwd() / "料金プラン.xlsx"

NAVY = "1F2937"
SOFT_GOLD = "FEF3C7"
LIGHT_GRAY = "F3F4F6"
WHITE = "FFFFFF"

YEN_FMT = '#,##0"円"'
USD_FMT = '"$"#,##0.00'
USD3_FMT = '"$"#,##0.000'  # $0.015 など小額単価用
PERSON_FMT = '#,##0"人"'
PCT_FMT = '0.0"%"'
GB_FMT = '0.00"GB"'

# シート名に全角括弧などの記号を含めると Numbers が数式参照を取り込めない
PLAN_SHEET = "料金プラン"
INFRA_SHEET = "インフラ詳細"

PLAN_REF = PLAN_SHEET  # Numbers 互換のためクォートしない

# TODO: 課金主体の呼称を案件に合わせる (例: ホテル / 配膳会社、企業 / 店舗)
ORG_LABEL = "管理組織"
TENANT_LABEL = "テナント"


def thin_border() -> Border:
    s = Side(style="thin", color="9CA3AF")
    return Border(left=s, right=s, top=s, bottom=s)


def fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=color, end_color=color)


def header_cell(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = fill(NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()


def section_title_cell(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, size=13, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def group_label_cell(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, size=11, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def phase_title_cell(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, size=14, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def note_cell(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(italic=True, color="6B7280", size=10)


def labeled_value_cell(ws: Worksheet, row: int, label: str, value, bg: str, fmt: str | None, note: str, note_col: int) -> None:
    ws.cell(row=row, column=1, value=label).border = thin_border()
    cell = ws.cell(row=row, column=2, value=value)
    cell.fill = fill(bg)
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt
    note_cell(ws.cell(row=row, column=note_col), note)


# ─── 料金プランシート (設定込み・営業向け) ──────────────────────────────


def build_plan(ws: Worksheet) -> dict[str, str]:
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 74

    phase_title_cell(ws.cell(row=1, column=1), "SaaS 料金プラン")  # TODO: サービス名を入れる
    note_cell(ws.cell(row=2, column=1), f"1 契約 = {ORG_LABEL} 1 + {TENANT_LABEL} (1 社〜) の月額課金。黄色のセルを編集すると全体が自動で再計算されます。")

    # ■ この料金の考え方
    section_title_cell(ws.cell(row=4, column=1), "■ この料金の考え方")
    ideas = [
        "1. 毎月必ずかかる費用 (エンジニアの運用費・インフラ) は「基本料」で回収する",
        "2. 人数に比例して増える費用は「従量課金」で回収する → 人が増えても赤字にならない",
        "3. 繁忙期は人数と一緒に収入も増える。閑散期は自動で安くなり、顧客にも公平",
    ]
    for i, text in enumerate(ideas):
        ws.cell(row=5 + i, column=1, value=text).font = Font(size=11)

    # ■ 料金表
    section_title_cell(ws.cell(row=9, column=1), "■ 料金表")
    for col, h in enumerate(["課金先", "月額", "単位", "含まれるもの"], start=1):
        header_cell(ws.cell(row=10, column=col), h)

    def fee_row(row: int, label: str, formula: str, unit: str, desc: str) -> None:
        ws.cell(row=row, column=1, value=label).border = thin_border()
        c = ws.cell(row=row, column=2, value=formula)
        c.border = thin_border()
        c.number_format = YEN_FMT
        c.font = Font(bold=True)
        c = ws.cell(row=row, column=3, value=unit)
        c.border = thin_border()
        c = ws.cell(row=row, column=4, value=desc)
        c.border = thin_border()
        c.alignment = Alignment(wrap_text=True, vertical="center")

    # TODO: 含まれるものの記述を案件の機能に合わせる
    fee_row(11, ORG_LABEL, "=$B$25", "/月", "解析ダッシュボード・帳票出力・データ長期保存 (R2 アーカイブ込み)")
    fee_row(12, TENANT_LABEL, "=$B$26", "/月/社", "利用・承認・アカウント管理。基本料に含まれる人数まで込み")
    fee_row(13, "従量 (超過分のみ)", "=$B$27", "/月/人", "含まれる人数を超えたアクティブユーザー 1 人あたり。アクティブ = その月に 1 回以上利用した人")

    # ■ 前提と設定
    section_title_cell(ws.cell(row=15, column=1), "■ 前提と設定（黄色のセル = 編集できます）")

    def input_cell(row: int, label: str, value, fmt: str | None = None, note: str = "") -> None:
        labeled_value_cell(ws, row, label, value, SOFT_GOLD, fmt, note, note_col=4)

    def calc_cell(row: int, label: str, formula: str, fmt: str | None = None, note: str = "") -> None:
        labeled_value_cell(ws, row, label, formula, LIGHT_GRAY, fmt, note, note_col=4)

    # TODO: 既定値をヒアリング結果に合わせる
    group_label_cell(ws.cell(row=16, column=1), "【契約構成】")
    input_cell(17, f"{ORG_LABEL}数", 1, note="解析ダッシュボードを使う組織の数")
    input_cell(18, f"{TENANT_LABEL}数", 1, note="利用するテナントの数。1 社でも赤字にならない料金設計にする")
    input_cell(19, "平均アクティブユーザー (人/月/社)", 100, PERSON_FMT, note="その月に 1 回以上利用した人数の平均 (1 社あたり)")
    input_cell(20, "繁忙期アクティブユーザー (人/月/社)", 400, PERSON_FMT, note="繁忙月の利用人数 (1 社あたり)")
    input_cell(21, "繁忙月数 (月/年)", 4, note="1 年のうち繁忙期にあたる月数")
    calc_cell(22, "平均月の合計人数 (人)", "=B19*B18", PERSON_FMT, note="計算: 平均アクティブ × テナント数")
    calc_cell(23, "繁忙月の合計人数 (人)", "=B20*B18", PERSON_FMT, note="計算: 繁忙期アクティブ × テナント数")

    group_label_cell(ws.cell(row=24, column=1), "【料金】")
    input_cell(25, f"{ORG_LABEL}基本料 (円/月)", 60000, YEN_FMT, note="解析ダッシュボード・帳票出力・長期保存。最小構成 (テナント 1 社・従量ゼロ) で固定費を上回るように設定する")
    input_cell(26, f"{TENANT_LABEL}基本料 (円/月/社)", 50000, YEN_FMT, note="利用・承認・アカウント管理")
    input_cell(27, "従量単価 (円/月/人)", 150, YEN_FMT, note="基本料に含まれる人数を超えた分だけ課金")
    input_cell(28, "基本料に含まれる人数 (人/社)", 100, PERSON_FMT, note="この人数まで基本料に込み。超えた分が従量")

    group_label_cell(ws.cell(row=29, column=1), "【人件費（毎月の運用）】")
    input_cell(30, "エンジニア単価 (円/h)", 5000, YEN_FMT, note="← ここに時間単価を入力")
    input_cell(31, "月あたり稼働時間 (h/月)", 20, note="監視・障害対応・問い合わせなどの運用時間")
    calc_cell(32, "月次人件費 (円/月)", "=B30*B31", YEN_FMT, note="計算: 単価 × 稼働時間 (自動算出)")

    group_label_cell(ws.cell(row=33, column=1), "【インフラ（外部サービス）】")
    input_cell(34, "為替 (円/USD)", 155, note="Cloudflare / AWS の請求はドル建て")
    input_cell(35, "Cloudflare 有料プラン基本料 ($/月)", 5, USD_FMT, note="Workers Paid プラン")
    input_cell(36, "認証 無料枠 (人/月)", 10000, PERSON_FMT, note="AWS Cognito Essentials の無料 MAU 枠")
    input_cell(37, "認証 超過単価 ($/人)", 0.015, USD3_FMT, note="無料枠を超えたログインユーザー 1 人あたり")
    input_cell(38, "管理ユーザー数 (人)", 20, PERSON_FMT, note="管理者・責任者など毎月ログインする人数")
    input_cell(39, "ログインユーザー (人/月/社)", 300, PERSON_FMT, note="月内にログインする一般ユーザー数 (1 社あたり、上限側の保守値)")
    input_cell(40, "D1 ストレージ込み枠 (GB)", 5, GB_FMT, note="Workers Paid に含まれる D1 ストレージ")
    input_cell(41, "D1 ストレージ超過単価 ($/GB/月)", 0.75, USD3_FMT, note="込み枠を超えた分。1 DB の上限は 10GB")
    input_cell(42, "R2 無料枠 (GB)", 10, GB_FMT, note="長期アーカイブ置き場。取り出しも無料")
    input_cell(43, "R2 超過単価 ($/GB/月)", 0.015, USD3_FMT, note="無料枠を超えた分。D1 の 1/50 の単価")

    # ■ 月次収支シミュレーション
    section_title_cell(ws.cell(row=45, column=1), "■ 月次収支シミュレーション")
    for col, h in enumerate(["項目", "平均月", "繁忙月", "計算式"], start=1):
        header_cell(ws.cell(row=46, column=col), h)

    def sim_row(row: int, label: str, avg_formula: str, peak_formula: str, memo: str, *, accent: bool = False, fmt: str = YEN_FMT) -> None:
        c = ws.cell(row=row, column=1, value=label)
        c.border = thin_border()
        if accent:
            c.font = Font(bold=True)
        for col, formula in ((2, avg_formula), (3, peak_formula)):
            cell = ws.cell(row=row, column=col, value=formula)
            cell.border = thin_border()
            cell.number_format = fmt
            if accent:
                cell.font = Font(bold=True)
                cell.fill = fill(SOFT_GOLD)
        note_cell(ws.cell(row=row, column=4), memo)

    usage_rev_avg = "=MAX(0,$B$22-$B$28*$B$18)*$B$27"
    usage_rev_peak = "=MAX(0,$B$23-$B$28*$B$18)*$B$27"
    infra_ref = INFRA_SHEET  # Numbers 互換のためクォートしない

    sim_row(47, f"収入: {ORG_LABEL}基本料", "=$B$25*$B$17", "=$B$25*$B$17", f"{ORG_LABEL}基本料 × {ORG_LABEL}数")
    sim_row(48, f"収入: {TENANT_LABEL}基本料", "=$B$26*$B$18", "=$B$26*$B$18", f"{TENANT_LABEL}基本料 × {TENANT_LABEL}数")
    sim_row(49, "収入: 従量", usage_rev_avg, usage_rev_peak, "含まれる人数を超えた分 × 従量単価")
    sim_row(50, "収入合計", "=B47+B48+B49", "=C47+C48+C49", "", accent=True)
    sim_row(51, "コスト: 人件費", "=$B$32", "=$B$32", "エンジニア単価 × 月稼働時間")
    infra_core = f"=({infra_ref}!B14+{infra_ref}!B34)*$B$34"
    r2_yen = f"={infra_ref}!B35*$B$34"
    sim_row(52, "コスト: インフラ費 (認証 + サーバ/DB)", infra_core, infra_core, "認証 + Workers/D1 (インフラ詳細シート)")
    sim_row(53, "コスト: R2 アーカイブ (長期保存)", r2_yen, r2_yen, "R2 無料枠を超えた分を自動計算 (単価は上の設定)")
    sim_row(54, "コスト合計", "=B51+B52+B53", "=C51+C52+C53", "", accent=True)
    sim_row(55, "利益", "=B50-B54", "=C50-C54", "収入合計 − コスト合計", accent=True)
    sim_row(56, "利益率", "=IFERROR(B55/B50*100,0)", "=IFERROR(C55/C50*100,0)", "", fmt=PCT_FMT)

    # ■ 年間収支
    section_title_cell(ws.cell(row=58, column=1), "■ 年間収支")
    for col, h in enumerate(["項目", "年間", "", "計算式"], start=1):
        header_cell(ws.cell(row=59, column=col), h)

    def year_row(row: int, label: str, formula: str, memo: str, *, accent: bool = False, fmt: str = YEN_FMT) -> None:
        c = ws.cell(row=row, column=1, value=label)
        c.border = thin_border()
        if accent:
            c.font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.border = thin_border()
        cell.number_format = fmt
        if accent:
            cell.font = Font(bold=True)
            cell.fill = fill(SOFT_GOLD)
        ws.cell(row=row, column=3).border = thin_border()
        note_cell(ws.cell(row=row, column=4), memo)

    year_row(60, "年間収入", "=B50*(12-$B$21)+C50*$B$21", "平均月 × (12 − 繁忙月数) + 繁忙月 × 繁忙月数")
    year_row(61, "年間コスト", "=B54*(12-$B$21)+C54*$B$21", "同上")
    year_row(62, "年間利益", "=B60-B61", "", accent=True)
    year_row(63, "年間利益率", "=IFERROR(B62/B60*100,0)", "", fmt=PCT_FMT)

    # ■ 備考
    section_title_cell(ws.cell(row=65, column=1), "■ 備考")
    notes = [
        f"・毎月かかる費用のほとんどは 1 契約目で回収済みのため、{TENANT_LABEL}・{ORG_LABEL}の追加分はほぼ全額が利益になる",
        "・「アクティブユーザー」= その月に 1 回以上利用した人。登録しただけで利用していない人には課金しない",
        f"・{TENANT_LABEL}基本料にはデータ長期保存のストレージ原価を織り込み済み (金額は「インフラ詳細」シートで自動計算)",
        "・インフラ費の内訳と根拠は「インフラ詳細」シート参照",
    ]
    for i, text in enumerate(notes):
        ws.cell(row=66 + i, column=1, value=text).font = Font(size=10)

    # クロスシート参照用のセル番地 (相対参照。Numbers 互換のため $ を付けない)
    return {
        "ORGS": "B17",
        "TENANTS": "B18",
        "AVG_PER_CO": "B19",
        "PEAK_PER_CO": "B20",
        "PEAK_MONTHS": "B21",
        "AVG_MAU": "B22",
        "PEAK_MAU": "B23",
        "LABOR_MONTHLY": "B32",
        "FX": "B34",
        "CF_BASE": "B35",
        "AUTH_FREE_MAU": "B36",
        "AUTH_OVERAGE": "B37",
        "ADMIN_MAU": "B38",
        "LOGIN_PER_CO": "B39",
        "D1_INCLUDED": "B40",
        "D1_OVERAGE": "B41",
        "R2_FREE": "B42",
        "R2_OVERAGE": "B43",
    }


# ─── インフラ詳細シート (エンジニア向け) ──────────────────────────────


def build_infra(ws: Worksheet, p: dict[str, str]) -> None:
    S = PLAN_REF

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 70

    phase_title_cell(ws.cell(row=1, column=1), "インフラ詳細 (エンジニア向け)")
    note_cell(ws.cell(row=2, column=1), "Cloudflare Workers/D1/R2 + AWS Cognito の料金条件と使用量モデル。黄色セルは編集可。価格は公式ページで裏取りした日付を記載すること")  # TODO: 価格確認日と出典を記載
    note_cell(ws.cell(row=3, column=1), "前提: D1 は 1 DB あたり 10GB 上限。重い帳票生成はブラウザ側で行い Workers CPU を消費しない")

    def input_cell(row: int, label: str, value, fmt: str | None = None, note: str = "") -> None:
        labeled_value_cell(ws, row, label, value, SOFT_GOLD, fmt, note, note_col=5)

    def calc_cell(row: int, label: str, formula: str, fmt: str | None = None, note: str = "") -> None:
        labeled_value_cell(ws, row, label, formula, LIGHT_GRAY, fmt, note, note_col=5)

    # ■ 使用量モデル
    # TODO: 案件の機能に合わせて既定値を調整し、実装後は実測値に更新する
    section_title_cell(ws.cell(row=5, column=1), "■ 使用量モデル (1 ユーザあたり)")
    input_cell(6, "API リクエスト (件/人/月)", 600, note="主要操作 + 閲覧 + 承認の想定")
    input_cell(7, "D1 書込 (行/人/月)", 1000, note="業務レコード + 監査ログ")
    input_cell(8, "D1 読取 (行/人/月)", 50000, note="ダッシュボードのスキャンを人数按分した想定")
    input_cell(9, "データ蓄積 (MB/人/年)", 1.8, "0.0", note="レコード + 監査ログの増分")
    input_cell(10, "保存年数 (年)", 5, note="法定保存期間に合わせる")

    # ■ 認証 (AWS Cognito)
    section_title_cell(ws.cell(row=12, column=1), "■ 認証 (AWS Cognito Essentials・東京リージョン)")
    calc_cell(13, "認証 MAU (人/月)", f"={S}!{p['ADMIN_MAU']}+{S}!{p['LOGIN_PER_CO']}*{S}!{p['TENANTS']}", PERSON_FMT, note="計算: 管理ユーザー + ログインユーザー × テナント数")
    calc_cell(14, "認証 月額 ($)", f"=MAX(0,B13-{S}!{p['AUTH_FREE_MAU']})*{S}!{p['AUTH_OVERAGE']}", USD_FMT, note="計算: 無料枠超過分 × 超過単価。ティア段差がない単価従量型")
    note_cell(ws.cell(row=15, column=5), "TOTP (認証アプリ) の MFA は全プラン無料。SMS MFA は SNS 別課金のため使わない")

    # ■ D1 ストレージ収容設計
    section_title_cell(ws.cell(row=17, column=1), "■ D1 ストレージ収容設計 (1 DB = 10GB 上限)")
    input_cell(18, "DB 使用率の上限 (%)", 80, note="10GB を使い切らず余裕を残す割合。この範囲に収まるテナント数だけ同居させる")
    calc_cell(19, "1 社あたり稼働 (人月/年)", f"={S}!{p['AVG_PER_CO']}*(12-{S}!{p['PEAK_MONTHS']})+{S}!{p['PEAK_PER_CO']}*{S}!{p['PEAK_MONTHS']}", "#,##0", note="計算: 平均アクティブ × 通常月 + 繁忙期アクティブ × 繁忙月。蓄積は利用した月の分だけ増える")
    calc_cell(20, "1 社あたり蓄積 (GB)", "=B19*(B9/12)*B10/1024", GB_FMT, note="計算: 人月/年 × MB/人月 × 保存年数")
    calc_cell(21, "収容できるテナント数 (社)", "=ROUNDDOWN(10*B18/100/B20,0)", "#,##0", note="計算: 10GB × 使用率上限 ÷ 1 社あたり蓄積。超える場合は DB 分割 or R2 アーカイブ")
    calc_cell(22, "ストレージ原価 (円/社/月)", f"=(10-{S}!{p['D1_INCLUDED']})*{S}!{p['D1_OVERAGE']}*{S}!{p['FX']}/B21", YEN_FMT, note="計算: 10GB 時の D1 超過費を収容テナント数で分配。テナント基本料に織り込む")

    # ▼ ユーザー数別の感度分析
    note_cell(ws.cell(row=24, column=1), "▼ 早見表: 平均アクティブ (黄色セル) を書き換えると、その規模での収容テナント数とストレージ原価が計算されます")
    for col, h in enumerate(["平均アクティブ (人/月/社)", "稼働 (人月/年)", "蓄積 (GB)", "収容テナント数 (社)", "ストレージ原価 (円/社/月)"], start=1):
        header_cell(ws.cell(row=25, column=col), h)
    for i, avg in enumerate([50, 100, 150, 200, 300]):
        row = 26 + i
        c = ws.cell(row=row, column=1, value=avg)
        c.fill = fill(SOFT_GOLD)
        c.border = thin_border()
        c.number_format = PERSON_FMT
        c = ws.cell(row=row, column=2, value=f"=A{row}*(12-{S}!{p['PEAK_MONTHS']})+{S}!{p['PEAK_PER_CO']}*{S}!{p['PEAK_MONTHS']}")
        c.border = thin_border()
        c.number_format = "#,##0"
        c = ws.cell(row=row, column=3, value=f"=B{row}*($B$9/12)*$B$10/1024")
        c.border = thin_border()
        c.number_format = GB_FMT
        c = ws.cell(row=row, column=4, value=f"=ROUNDDOWN(10*$B$18/100/C{row},0)")
        c.border = thin_border()
        c.number_format = "#,##0"
        c = ws.cell(row=row, column=5, value=f"=(10-{S}!{p['D1_INCLUDED']})*{S}!{p['D1_OVERAGE']}*{S}!{p['FX']}/D{row}")
        c.border = thin_border()
        c.number_format = YEN_FMT

    # ■ Cloudflare
    section_title_cell(ws.cell(row=32, column=1), "■ Cloudflare (Workers Paid + D1 + R2)")
    calc_cell(33, "データ蓄積合計 (GB)", f"=B20*{S}!{p['TENANTS']}", GB_FMT, note="計算: 1 社あたり蓄積 × テナント数 (繁忙期の蓄積も含む人月ベース)")
    calc_cell(34, "Workers + D1 月額 ($)", f"={S}!{p['CF_BASE']}+MAX(0,B33-{S}!{p['D1_INCLUDED']})*{S}!{p['D1_OVERAGE']}", USD_FMT, note="計算: 基本料 + D1 ストレージ込み枠の超過分 × 超過単価")
    calc_cell(35, "R2 アーカイブ月額 ($)", f"=MAX(0,B33-{S}!{p['R2_FREE']})*{S}!{p['R2_OVERAGE']}", USD_FMT, note="D1 のバックアップ (Time Travel) は 30 日まで → 長期保存分を R2 へ日次エクスポート。取り出しも無料")

    # ▼ テナント数別の感度分析
    note_cell(ws.cell(row=37, column=1), "▼ 早見表: テナント数 (黄色セル) を書き換えると、その社数でのストレージ費 (D1 + R2) が計算されます")
    for col, h in enumerate(["テナント数 (社)", "蓄積合計 (GB)", "D1 超過費 ($/月)", "R2 費 ($/月)", "ストレージ費 (円/社/月)"], start=1):
        header_cell(ws.cell(row=38, column=col), h)
    for i, n_co in enumerate([5, 10, 20, 50, 100]):
        row = 39 + i
        c = ws.cell(row=row, column=1, value=n_co)
        c.fill = fill(SOFT_GOLD)
        c.border = thin_border()
        c.number_format = "#,##0"
        c = ws.cell(row=row, column=2, value=f"=$B$20*A{row}")
        c.border = thin_border()
        c.number_format = GB_FMT
        c = ws.cell(row=row, column=3, value=f"=MAX(0,B{row}-{S}!{p['D1_INCLUDED']})*{S}!{p['D1_OVERAGE']}")
        c.border = thin_border()
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=4, value=f"=MAX(0,B{row}-{S}!{p['R2_FREE']})*{S}!{p['R2_OVERAGE']}")
        c.border = thin_border()
        c.number_format = USD_FMT
        c = ws.cell(row=row, column=5, value=f"=(C{row}+D{row})*{S}!{p['FX']}/A{row}")
        c.border = thin_border()
        c.number_format = YEN_FMT

    # ■ 無料枠判定
    section_title_cell(ws.cell(row=45, column=1), "■ 無料枠 (Workers Free) 判定 — 参考")
    for col, h in enumerate(["項目", "無料枠上限 (/日)", "平均月の使用量 (/日)", "判定", "備考"], start=1):
        header_cell(ws.cell(row=46, column=col), h)
    free_rows = [
        ("API リクエスト", 100000, f"={S}!{p['AVG_MAU']}*$B$6/30", "主要操作 + 閲覧のみなら余裕"),
        ("D1 書込 (行)", 100000, f"={S}!{p['AVG_MAU']}*$B$7/30", "繁忙期でも枠内か確認"),
        ("D1 読取 (行)", 5000000, f"={S}!{p['AVG_MAU']}*$B$8/30", "解析ダッシュボードが蓄積データを叩くと超過しやすい (実質の Free 制約)"),
        ("D1 ストレージ (GB)", 5, "=$B$33", "長期蓄積で上限に近づく (実質の Free 制約)"),
    ]
    for i, (label, limit, usage_formula, memo) in enumerate(free_rows):
        row = 47 + i
        ws.cell(row=row, column=1, value=label).border = thin_border()
        c = ws.cell(row=row, column=2, value=limit)
        c.border = thin_border()
        c.number_format = "#,##0"
        c = ws.cell(row=row, column=3, value=usage_formula)
        c.border = thin_border()
        c.number_format = "#,##0.00" if label.endswith("(GB)") else "#,##0"
        c = ws.cell(row=row, column=4, value=f'=IF(C{row}<=B{row},"枠内","超過")')
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center")
        note_cell(ws.cell(row=row, column=5), memo)

    # ■ 月額インフラ費まとめ (料金プランシートは B14/B34/B35 を直接参照する)
    section_title_cell(ws.cell(row=52, column=1), "■ 月額インフラ費まとめ (円)")
    calc_cell(53, "インフラ費 平均月 (円)", f"=(B14+B34+B35)*{S}!{p['FX']}", YEN_FMT, note="計算: (認証 + Workers/D1 + R2) × 為替")
    calc_cell(54, "インフラ費 繁忙月 (円)", f"=(B14+B34+B35)*{S}!{p['FX']}", YEN_FMT, note="認証 MAU が繁忙期に増える場合はここを分けて計算する")
    calc_cell(55, "一人あたりインフラ費 (円/人/月)", f"=B53/{S}!{p['AVG_MAU']}", YEN_FMT, note="平均月ベース。従量単価の原価にあたる")


def main() -> None:
    wb = Workbook()
    plan_ws = wb.active
    plan_ws.title = PLAN_SHEET
    infra_ws = wb.create_sheet(INFRA_SHEET)

    p = build_plan(plan_ws)
    build_infra(infra_ws, p)

    # openpyxl は計算結果のキャッシュを書かないため、開いた時に全再計算させる (Numbers 互換に必須)
    wb.calculation.fullCalcOnLoad = True

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"generated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
